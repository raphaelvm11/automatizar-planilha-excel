import openpyxl

#criar uma planilha(pasta)
book = openpyxl.Workbook()
#vizualizar paginas existentes
print(book.sheetnames)
#como criar uma pagina
book.create_sheet('frutas')
#como selecionar uma pagina
frutas_page = book['frutas']
frutas_page.append(['FRUTA', 'QUANTIDADE','PREÇO'])
frutas_page.append(['banana','5','3,90'])
frutas_page.append(['maçã','3','2,80'])
frutas_page.append(['morango','4','4,30'])
frutas_page.append(['kiwi','7','4,60'])
#salvar a planilia
book.save('planilha de compras.xlsx')

